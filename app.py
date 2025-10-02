import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from PIL import Image
import os

# Load Excel workbook
excel_path = "Maintenance Tracker.xlsx"
wb = load_workbook(excel_path)
sheet_names = wb.sheetnames

# Load logo
logo_path = "logo.png"
if os.path.exists(logo_path):
    logo = Image.open(logo_path)
    st.sidebar.image(logo, use_column_width=True)

# User credentials
USER_CREDENTIALS = {
    "admin": {"password": "admin123", "role": "Supervisor"},
    "user": {"password": "user123", "role": "Technician"}
}

# Login
def login():
    st.sidebar.title("🔐 Login")
    username = st.sidebar.text_input("Username")
    password = st.sidebar.text_input("Password", type="password")
    if st.sidebar.button("Login"):
        user = USER_CREDENTIALS.get(username)
        if user and user["password"] == password:
            st.session_state.auth = True
            st.session_state.user = username
            st.session_state.role = user["role"]
        else:
            st.sidebar.error("Invalid credentials")

if "auth" not in st.session_state:
    st.session_state.auth = False
    login()

# Main app
if st.session_state.auth:
    st.title("🛠️ Equipment Maintenance Tracker")
    st.caption(f"Logged in as: {st.session_state.user} ({st.session_state.role})")

    tab_home, tab_browse, tab_detail = st.tabs(["🏠 Home", "📂 Browse", "🔍 Details"])

    with tab_home:
        st.subheader("Search by Tag")
        selected_sheet = st.selectbox("Select Area", sheet_names)
        sheet = wb[selected_sheet]
        tags = [cell.value for cell in sheet['A'][1:] if cell.value]
        selected_tag = st.selectbox("Select Tag", tags)

        if selected_tag:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == selected_tag:
                    st.write("**Equipment Details:**")
                    for i, cell in enumerate(row):
                        st.write(f"{sheet.cell(row=1, column=i+1).value}: {cell}")
                    break

    with tab_browse:
        st.subheader("Browse Equipment")
        selected_sheet = st.selectbox("Select Area to Browse", sheet_names, key="browse")
        sheet = wb[selected_sheet]
        data = [[cell.value for cell in row] for row in sheet.iter_rows(values_only=True)]
        df = pd.DataFrame(data[1:], columns=data[0])
        st.dataframe(df)

    with tab_detail:
        st.subheader("Equipment Details")
        selected_sheet = st.selectbox("Select Area for Details", sheet_names, key="detail")
        sheet = wb[selected_sheet]
        tags = [cell.value for cell in sheet['A'][1:] if cell.value]
        selected_tag = st.selectbox("Select Tag for Details", tags, key="detail_tag")

        if selected_tag:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == selected_tag:
                    st.write("**Details:**")
                    for i, cell in enumerate(row):
                        st.write(f"{sheet.cell(row=1, column=i+1).value}: {cell}")
                    break

            if st.session_state.role == "Supervisor":
                st.subheader("✏️ Edit Equipment")
                edited_values = []
                for i, cell in enumerate(row):
                    new_val = st.text_input(f"{sheet.cell(row=1, column=i+1).value}", value=cell)
                    edited_values.append(new_val)

                if st.button("Save Changes"):
                    for i, val in enumerate(edited_values):
                        sheet.cell(row=row[0]+2, column=i+1).value = val
                    wb.save(excel_path)
                    st.success("Changes saved successfully.")
            else:
                st.info("You are signed in as a Technician and cannot edit records.")

    # Admin-only logo upload
    if st.session_state.role == "Supervisor":
        st.sidebar.subheader("📤 Upload Logo")
        uploaded_logo = st.sidebar.file_uploader("Upload PNG logo", type=["png"])
        if uploaded_logo:
            with open("logo.png", "wb") as f:
                f.write(uploaded_logo.getbuffer())
            st.sidebar.success("Logo updated. Please refresh the app.")

else:
    st.warning("Please log in to access the app.")